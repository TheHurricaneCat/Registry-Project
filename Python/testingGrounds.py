import threading
import sqlite3 as lite
import kivy 
from kivy.app import App 
from kivy.uix.label import Label
from kivy.uix.gridlayout import GridLayout
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.widget import Widget
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.floatlayout import FloatLayout

from kivy.uix.scrollview import ScrollView
from kivy.core.window import Window
from kivy.uix.popup import Popup

import random

import openpyxl
from datetime import date

global prodInfo
prodInfo = ["Staff Name"]

global defaultNumMsg
defaultNumMsg = "ENTER THE QUANTITY"

global defaultProdMsg
defaultProdMsg = "ENTER THE PRODUCT"

global staffFile
staffFile = "C:/Users/Josefe Gillego/Documents/Special Project/Python/staff.txt"

with open(staffFile) as f:
    global staffList
    staffList = f.read().splitlines()
    f.close()

#Classes for the product slice
class ProdMod(Label):
    pass

class ProdScroll(BoxLayout):
    pass

class ProdSelection(BoxLayout):
    global curProd
    curProd = defaultProdMsg
    def onProdPress(self):
        global curProd
        curProd = self.ids.button1.text
        updateProd()
class CustomButton(Button):
    pass

#Classes for the quantity slice
class KeyContainer(BoxLayout):
    pass

class KeyMod(Label):
    pass

class KeyPad(BoxLayout):
    global curNum
    curNum = defaultNumMsg
    def onKeyPress(self):
        global curNum
        global curProd
        global prodInfo
        if self.ids.button2.text == "DEL":
            curNum = defaultNumMsg
            updateKeys()
            print("Status: deleting quantity")
        elif self.ids.button2.text == "ENTER":
            if (curNum != defaultNumMsg and curProd != defaultProdMsg and int(curNum)*2 != 0):
                prodInfo = prodInfo + [str(curProd)] + [str(curNum)]
                updateView()
                curNum = defaultNumMsg
                curProd = defaultProdMsg
                updateKeys()
                updateProd()
            else: 
                if (curNum == defaultNumMsg and curProd == defaultProdMsg): 
                    error = StaffError()
                    error.open()
                    error.title = "Error: No product or quantity chosen"
                    print("Error: quantity is zero or not stated ")
                elif (curNum == defaultNumMsg or int(curNum)*2 == 0): 
                    error = StaffError()
                    error.open()
                    error.title = "Error: No quantity entered or is zero"
                    print("Error: quantity is zero")
                elif (curProd == defaultProdMsg): 
                    error = StaffError()
                    error.open()
                    error.title = "Error: No product chosen"
                    print("Error: product is not stated ")
        else:
            if (curNum == "Default" or curNum == defaultNumMsg):
                curNum = ""
            curNum = curNum + str(self.ids.button2.text)
            updateKeys()

#Classes for overview slice
class ViewContainer(BoxLayout):
    pass

class ViewMod(Label):
    pass

class ViewScroll(BoxLayout):
    pass

class ViewPad(BoxLayout):
    def onClearPress(self):
        delElem()
    def onEnterPress(self):
        element = [i for i in addViewScroll.ids.viewScrollContainer.children]
        print(element)
        if (currStaff != "" and len(element) >= 1):
            processData()
        elif currStaff == "":
            error = StaffError()
            error.open()
            error.title = "Error: No staff selected"
        elif (len(element) < 1):
            error = StaffError()
            error.open()
            error.title = "Error: No item in the catalog"
        
class ViewSelection(BoxLayout):
    global index 
    index = 1
    def removeItem(self):
        global prodInfo
        prodInfo[self.index+1] = ""
        prodInfo[self.index] = ""
        self.parent.remove_widget(self)

#Classes for staff slice
class StaffPad(BoxLayout):
    def selectStaff(self):
        print("deselecting")
    def confirmStaff(self):
        updateCurrentStaff()
        
class StaffSelection(BoxLayout):
    global currStaff
    currStaff = ""
    def updateStaff(self):
        global currStaff
        currStaff = self.ids.staffTitle.text
        staffEntry.ids.currentStaff.text = currStaff
        confirmation = StaffConfirm()
        confirmation.open()
        confirmation.title = "Set <" + currStaff + "> as the current staff?"
        
class StaffConfirmation(Popup):
    def __init__(self, **kwargs):
        super(StaffConfirmation, self).__init__(**kwargs)
        for person in staffList:
            staffEntry = StaffSelection()
            staffEntry.ids.staffTitle.text = str(person)
            self.ids.staffContainer.add_widget(staffEntry)
    def confirmStaff(self):
        staffEntry = StaffSelection()
        global currInput
        currInput = self.ids.textInput.text
        if (currInput != ""):
            staffEntry.ids.staffTitle.text = currInput
            self.ids.staffContainer.add_widget(staffEntry)
            self.ids.textInput.text = currInput + " is successully added!"
            f = open(staffFile,"a+")
            f.write("\n")
            f.write(currInput)
            f.close()
        else:
            error = StaffError()
            error.open()
            error.title = "Error: No input"
    def clearField(self):
        self.ids.textInput.text = ""

class StaffError(Popup):
    pass

class StaffConfirm(Popup):
    def setStaff(self):
        self.dismiss()
        staffEntry.dismiss()
        updateCurrentStaff()

#Main Grid
class TestGrid(Widget):
    def __init__(self, **kwargs):
        super(TestGrid, self).__init__(**kwargs)
        
        #Color values
        #Window Color
        parentBgColor = 249/255, 246/255, 238/255, 1 

        #Product selection color control
        prodTitleColor = 39/255, 41/255, 54/255, 1 #Blueish gray
        prodScrollColor = 154/255, 157/255, 157/255, 1 #Whiteish gray
        prodItemColor = 154/255,  157/255, 157/255, 1
        
        keyTitleColor = 39/255, 41/255, 54/255, 1

        viewTitleColor = 39/255, 41/255, 54/255, 1

        self.ids.bgGrid.background_color = parentBgColor
        #Generate the Product List 
        
        prodTitle = ProdMod()
        self.ids.prodGrid.add_widget(prodTitle)
        prodTitle.background_color = prodTitleColor 

        addProdScroll = ProdScroll()
        self.ids.prodGrid.add_widget(addProdScroll)
        addProdScroll.background_color = prodScrollColor
        
        products =  ["Yakuza Teriyaki", "Chicano Chili", "Waddup Che&Bac","Gangbanger Tuna", "Rastaparay Veg", "Hardcore Overload"]
        counter = 1
        for num in products:
            productButton = ProdSelection()
            productButton.ids.varcont = num
            productButton.ids.button1.text = num
            addProdScroll.ids.prodContainer.add_widget(productButton)
            productButton.background_color = prodItemColor
            productButton.ids.prodImage.source = 'RegistryImages/' + str(counter) + ".png"
            if (counter % 3 == 0):
                productButton.ids.label1.background_color = 71/255, 102/255, 194/255, 1
            elif (counter % 2 == 0):
                productButton.ids.label1.background_color = 61/255, 205/255, 196/255, 1
            else:
                productButton.ids.label1.background_color = 255/255, 222/255, 89/255, 1
            counter+=1
            print(counter)

        #Generate the Keypad
        keyTitle = KeyMod()
        self.ids.keyGrid.add_widget(keyTitle)
        keyTitle.background_color = keyTitleColor

        addKeys = KeyContainer()
        self.ids.keyGrid.add_widget(addKeys)
        
        keys = [7, 8, 9, 4, 5, 6, 1, 2, 3,"ENTER", 0, "DEL"]
        for num in keys:
            keyButton = KeyPad()
            keyButton.ids.button2.text = str(num)
            if (num == "ENTER"):
                keyButton.ids.button2.background_color = 111/255, 159/255, 91/255, 1
            elif (num == "DEL"):
                keyButton.ids.button2.background_color = 198/255, 92/255, 63/255, 1
            else:
                keyButton.ids.button2.background_color = 43/255, 50/255, 58/255, 1
            addKeys.ids.keyContainer.add_widget(keyButton)
            

        #Generate the overview

        viewTitle = ViewMod()
        self.ids.viewGrid.add_widget(viewTitle)
        viewTitle.background_color = viewTitleColor
        
        global addViewScroll
        addViewScroll = ViewScroll()
        self.ids.viewGrid.add_widget(addViewScroll)
        
        addConfirm = ViewContainer()
        self.ids.viewGrid.add_widget(addConfirm)
        
        confirmButton = ViewPad()
        addConfirm.ids.viewContainer.add_widget(confirmButton)

        #Update changes
        global updateProd
        global updateKeys
        global updateView
        global updateCurrentStaff
        global delElem
        global processData
        def updateProd():
            prodTitle.text = str(curProd)
        def updateKeys():
            keyTitle.text = str(curNum)
        def updateView():
            global index
            addProductView = ViewSelection()
            addProductView.ids.viewTitle.text = prodInfo[-1-1]
            addProductView.ids.viewQuantity.text = str("Qty:") + " " + prodInfo[-1]
            addProductView.ids.viewImage.source = 'RegistryImages/' + str(products.index(addProductView.ids.viewTitle.text)+1) + ".png"
            addProductView.index = index
            color = products.index(addProductView.ids.viewTitle.text)+1
            #Add color seperators
            if (color % 3 == 0):
                addProductView.ids.viewIndc.background_color = 71/255, 102/255, 194/255, 1
            elif (color % 2 == 0):
                addProductView.ids.viewIndc.background_color = 61/255, 205/255, 196/255, 1
            else:
                addProductView.ids.viewIndc.background_color = 255/255, 222/255, 89/255, 1
            
            addViewScroll.ids.viewScrollContainer.add_widget(addProductView)
            index = index+2
            print(prodInfo)
        def updateCurrentStaff(): 
            print("updating")
            self.ids.staffTitle.text = "Current Staff: " + currStaff
            prodInfo[0] = currStaff
        def processData():
            spaces = 0
            for space in prodInfo:
                if (space == ""):
                    spaces+=1
            for item in range(spaces):
                prodInfo.remove("")
            excelProcess()
            delElem()
            error = StaffError()
            error.open()
            error.title = "Entry sucessfully registered"
        def delElem():
            row = [i for i in addViewScroll.ids.viewScrollContainer.children]
            for child in row:
                addViewScroll.ids.viewScrollContainer.remove_widget(child)
            prodInfo[1:] = ""
            global index
            index = 1
            print(prodInfo)
        def excelProcess():
            wb = openpyxl.load_workbook(r"C:/Users/Josefe Gillego/Documents/Special Project/Python/TestFile.xlsx")
            ws = wb["Sheet1"]

            # Get the column length and ignore whitespaced cells
            currentLocation = 1
            for i in range(1, ws.max_row):
                if ws["A" + str(i)].value != None:
                    print(currentLocation)
                    currentLocation += 1

            # Write data in excel
            colNames = ["A", "B", "C", "D", "E"]

            # Initialize invoice number
            invNum = str(ws["D" + str(currentLocation - 1)].value)
            print(invNum)
            print(prodInfo)

            for column in range(5):
                # writing the date, staff and invoice number
                def writeExcel(identifier, data):
                    ws[str(colNames[column]) + str(currentLocation + item - identifier)] = data
                
                for item in range(1, int((len(prodInfo)) / 2) + 1):
                    if (column == 0): 
                        writeExcel(1, date.today())
                    if (column == 3): 
                        multi = False
                        if (multi == False):
                            writeExcel(1, int(date.today().strftime('%y%m%d') + format(int(invNum[6:]) + 1, '004d')))
                            multi = True
                        else:
                            writeExcel(1, int(ws[str(colNames[column]) + str(currentLocation - 1)]))
                    if (column == 4):
                            writeExcel(1, prodInfo[0])
                # write the product code and quantity
                identA = 1
                identB = 2
                for item in range(1, len(prodInfo)):
                    if (column == 1 and item%2 != 0):
                        writeExcel(identA, prodInfo[item])
                        identA += 1
                    if (column == 2 and item%2 == 0): 
                        writeExcel(identB, int(prodInfo[item]))
                        identB += 1

            wb.save(r"C:/Users/Josefe Gillego/Documents/Special Project/Python/TestFile.xlsx")
            print("excel written")

    def activatePopup(self):
        global staffEntry
        staffEntry = StaffConfirmation()
        staffEntry.open()
        self.ids.staffTitle.background_color = random.uniform(250/255, 255/255), 0, 0, 1
    
class TestApp(App):
    def build(self):
        Window.size = 1100, (4*120)
        return TestGrid()
    
if __name__ == '__main__':
    TestApp().run()
