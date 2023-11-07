import openpyxl
from datetime import date

wb = openpyxl.load_workbook(r"C:/Users/Josefe Gillego/Documents/Special Project/Python/TestFile.xlsx")
ws = wb["Sheet1"]

# Get the column length and ignore whitespaced cells
currentLocation = 1
for i in range(1, ws.max_row):
    if ws["A" + str(i)].value != None:
        currentLocation += 1

# Get user input
itmDetails = [""]
    
print("Enter the staff")
itmDetails[0] = input()
    
line = 0
while (True):
    print("Enter the product code: ")
    itmDetails = itmDetails + [input()]
    if (itmDetails[line+1] == "D"):
        break
    line += 1
    print("Enter the item quantity")
    itmDetails = itmDetails + [input()]
    line += 1

# Write data in excel
colNames = ["A", "B", "C", "D", "E"]

# Initialize invoice number
invNum = str(ws["D" + str(currentLocation - 1)].value)

for column in range(5):
    # writing the date, staff and invoice number
    def writeExcel(identifier, data):
        ws[str(colNames[column]) + str(currentLocation + item - identifier)] = data
    
    for item in range(1, int((len(itmDetails) - 2) / 2) + 1):
        if (column == 0): 
            writeExcel(1, date.today())
        if (column == 3): 
            multi = False
            if (multi == False):
                writeExcel(1, int(date.today().strftime('%y%m%d') + format(int(invNum[6:]) + 1, '004d')))
                multi = True
            else:
                writeExcel(1, int(ws[str(colNames[column]) + str(currentLocation - 1)]))
        if (column == 4):
                writeExcel(1, itmDetails[0])
    # write the product code and quantity
    identA = 1
    identB = 2
    for item in range(1, len(itmDetails) - 1):
        if (column == 1 and item%2 != 0):
            writeExcel(identA, itmDetails[item])
            identA += 1
        if (column == 2 and item%2 == 0): 
            writeExcel(identB, int(itmDetails[item]))
            identB += 1

wb.save(r"C:/Users/Josefe Gillego/Documents/Special Project/Python/TestFile.xlsx")

