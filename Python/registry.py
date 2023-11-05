#Based Code
import openpyxl
from datetime import date

wb = openpyxl.load_workbook(r"C:/Users/Josefe Gillego/Documents/Special Project/Python/TestFile.xlsx")
ws = wb["Sheet1"]

placeholder = ["Date", "Product Code", "Staff", "Invoice Number", "Quantity"]
inputMsg = ["", "", "", "", ""]

for msg in range(1, 5):
    if (msg != 3):
        print("Enter the " + placeholder[msg] + ":")
        inputMsg[msg] = input()

# Get the column length and ignore whitespaced cells
currentLocation = 0
for i in range(1, ws.max_row):
    if ws["A" + str(i)].value != None:
        currentLocation = i

# Initialize invoice number
invNum = str(ws["D" + str(currentLocation)].value)

# Writing the input data
columnSelection = ["A", "B", "C", "D", "E"]
for column in range(5):
    
    # write on the spreadsheet
    if column == 3:
        ws[str(columnSelection[column]) + str(currentLocation + 1)] = int(date.today().strftime('%y%m%d') + format(int(invNum[6:]) + 1, '004d'))
    elif column == 1 or column == 2: 
        ws[str(columnSelection[column]) + str(currentLocation + 1)] = inputMsg[column]
    elif column == 0:
        ws[str(columnSelection[column]) + str(currentLocation + 1)] = date.today()
    else:
        ws[str(columnSelection[column]) + str(currentLocation + 1)] = int(inputMsg[column])

wb.save(r"C:/Users/Josefe Gillego/Documents/Special Project/Python/TestFile.xlsx")