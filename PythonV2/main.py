import kivy
import openpyxl
from datetime import date
from kivy.app import App 
from kivy.uix.label import Label
from kivy.uix.gridlayout import GridLayout
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.widget import Widget
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.floatlayout import FloatLayout

from kivy.uix.scrollview import ScrollView
from kivy.core.window import Window
from kivy.uix.popup import Popup

class MainGrid(Widget):
    def __init__(self, **kwargs):
        super(MainGrid, self).__init__(**kwargs)

class MainApp(App):
    def build(self):
        Window.maximize()
        return MainGrid()
    
if __name__ == '__main__':
    MainApp().run()

def processExcel():
    wb = openpyxl.load_workbook(r"TestFile.xlsx")
    ws = wb["Sheet1"]

    prodList = ["Staff Name", "Product 1", 20, "Product 2", 30]

    # Get the column length and ignore whitespaced cells
    currentLocation = 1
    for i in range(1, ws.max_row):
        if ws["A" + str(i)].value != None:
            currentLocation += 1

    # Write data in excel
    colNames = ["A", "B", "C", "D", "E"]

    # Initialize invoice number
    invNum = str(ws["D" + str(currentLocation - 1)].value)

    for column in range(5):
        # writing the date, staff and invoice number
        def writeExcel(identifier, data):
            ws[str(colNames[column]) + str(currentLocation + item - identifier)] = data
        
        for item in range(1, int((len(prodList)) / 2) + 1):
            if (column == 0): 
                writeExcel(1, date.today())
            if (column == 3): 
                multi = False
                if (multi == False):
                    writeExcel(1, int(date.today().strftime('%y%m%d') + format(int(invNum[6:]) + 1, '004d')))
                    multi = True
                else:
                    writeExcel(1, int(ws[str(colNames[column]) + str(currentLocation - 1)]))
            if (column == 4):
                    writeExcel(1, prodList[0])
        # write the product code and quantity
        identA = 1
        identB = 2
        for item in range(1, len(prodList)):
            if (column == 1 and item%2 != 0):
                writeExcel(identA, prodList[item])
                identA += 1
            if (column == 2 and item%2 == 0): 
                writeExcel(identB, int(prodList[item]))
                identB += 1

    wb.save(r"TestFile.xlsx")
    print("excel written")